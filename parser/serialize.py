# import module
import openpyxl
from enum import Enum
from entity import Cell
import yaml

class EntityType(Enum):
   EXCEL = 1
   CELL = 2
   FUNCTION = 3
   PLOT = 4

def map_number_to_letter(number) -> str:
   '''
   1 -> A
   '''
   return chr(ord('@')+number)

def serialize_excel_yaml(filename):
# load excel with its path
   try:
      wrkbk = openpyxl.load_workbook(filename)
   except:
      return ValueError(f'{filename} is not a valid path to an excel file')
   
   sheet = wrkbk.active
   entities = []
   yaml_obj = {'Source': {'Type': 'EXCEL', 'Path': filename, 'Entities': entities}}

   # iterate through excel and display data
   for row in range(1, sheet.max_row+1):         
      for col in range(1, sheet.max_column+1):
         col_letter = map_number_to_letter(col)
         cell_obj = sheet.cell(row=row, column=col)
         entity = {'Entity': {'type': 'cell', 'value': cell_obj.value, 'location': f'{col_letter}{row}'}}
         entities.append(entity)
         # print(cell_obj.value, end=" ")

   with open(f'{filename}.yaml', 'w') as file:
      yaml.dump(yaml_obj, file)

def main():
   pass
   # filename = '/Users/sarboroy/Downloads/uscities.xlsx'
   # serialize_excel_yaml(filename)

if __name__ == "__main__":
   main()