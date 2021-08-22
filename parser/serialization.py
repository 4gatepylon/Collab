# import module
import openpyxl
from enum import Enum
import yaml

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE

import pprint as _pp
pp = _pp.PrettyPrinter(indent=4)
def pprint(string):
    pp.pprint(string)

# Yaml Declarations
__SOURCES = "Sources"
__SOURCE = "Source"
__TYPE = "Type"
__PATH = "Path"
__NAME = "Name"         # Optional
__ENTITIES = "Entities" # Optional
_ID = "Id"
# __NAME                # Optional
__LOC = "CellLocation"  # Optional
__CACHE = "CachedVal"   # Optional
__VAL = "FuncVal"       # Optional
__GROUPS = "Groups"     # Optional
__META = "Metadata"     # Optional
__CHILDREN = "Children"

# Source Type
__PYTHON_STDOUT_FILE = "PYTHON_STDOUT_FILE"
__EXCEL_FILE = "EXCEL_FILE"
########################################################### Deserialization
# def deserialize(filename):
#     wb = load_workbook(filename=filename)
#     yaml = {}
#     yaml["sheets"] = []
#     for sheet_name in wb.sheetnames:
#         sheet = deserialize_sheet(wb[sheet_name])
#         sheet["Type"] = "Sheet"
#         sheet["Name"] = sheet_name
#         yaml["sheets"].append(sheet)
#     return yaml

# # Note that coords start at 1 instead of 0
# # def cell_name2coords(cell_name):
# #     pass

# # do not support merged cells for now
# def deserialize_sheet(sheet):
#     yaml = {}
#     yaml["cells"] = []
#     # Max containing data
#     for column in range(sheet.min_column, sheet.max_column + 1):
#         for row in range(sheet.min_row, sheet.max_row + 1):
#             cell = sheet.cell(row, column)
#             if cell.value != None:
#                 yaml["cells"].append({
#                     "Type": "Cell",
#                     "Name": cell.coordinate,
#                     "FuncValue": cell.value,
#                 })
#     return yaml

num_letters = ord("Z") - ord("A") + 1

def letter2number(letter):
   if ord(letter) < ord("A") or ord(letter) > ord("Z"):
      raise Exception("Must be uppercase letter!")
   # A => 1, B => 2, ...
   return ord(letter) - ord("A") + 1

# TODO this might be buggy
def col_letter2col(col_letter):
   if len(col_letter) > 1:
      raise NotImplementedError
   acc = 0
   for let in col_letter:
      acc *= num_letters
      acc += letter2number(let)
   return acc

def deserialize_loc(loc):
   i = 0
   while i < len(loc) and not loc[i].isdigit():
      i += 1
   return col_letter2col(loc[:i]), int(loc[i:])

def deserialize(excel_path, yaml_path):
   wb = Workbook()
   y = None
   with open(yaml_path) as f:
      # load_all will give you a stream please ignore it
      y = yaml.load(f, yaml.loader.SafeLoader)
   ws = wb.create_sheet(title="sheet")
   # print(y)
   for e in y["Source"]["Entities"]:
      entity = e["Entity"]
      loc = entity["location"]
      t = entity["type"]
      val = entity["value"]
      if t == "cell":
         col, row = deserialize_loc(loc)
         _ = ws.cell(column=col, row=row, value=val)
      else:
         raise NotImplementedError
      # pprint(entity)
   wb.save(filename=excel_path)

########################################################### Serialization
class EntityType(Enum):
   EXCEL = 1
   CELL = 2
   FUNCTION = 3
   PLOT = 4

# TODO support
def number2letters(number) -> str:
   if number > num_letters:
      raise NotImplementedError
   return chr(ord('A') - 1 + number)

def serialize_excel_yaml(filename):
# load excel with its path
   try:
      wrkbk = openpyxl.load_workbook(filename)
   except:
      return ValueError(f'{filename} is not a valid path to an excel file')
   
   sheet = wrkbk.active
   entities = []
   yaml_obj = {'Source': {'Type': 'EXCEL', 'Path': filename, 'Entities': entities}}

   # iterate through excel and display data
   for row in range(1, sheet.max_row+1):         
      for col in range(1, sheet.max_column+1):
         col_letter = number2letters(col)
         cell_obj = sheet.cell(row=row, column=col)
         entity = {'Entity': {'type': 'cell', 'value': cell_obj.value, 'location': f'{col_letter}{row}'}}
         entities.append(entity)
         # print(cell_obj.value, end=" ")

   with open(f'{filename}.yaml', 'w') as file:
      yaml.dump(yaml_obj, file)

if __name__ == "__main__":
   # d = deserialize("ex.xlsx")
   deserialize("sample.xlsx", "sample.yaml")
   # pprint(d)
   # print("Serializing into test_output.yaml")
   # serialize_excel_yaml()