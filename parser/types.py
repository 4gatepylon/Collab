import uuid
import yaml
import openpyxl

# YAML Might look like:
# Source (list of these)
# - Type [EXCEL | PYTHON_STDOUT]
# - Path
# - Id
#
#
# Entity (list of these)
# - Type [Excel, Cell | Plot | Func | ... ]
# - LocStr  [| None]
# - FuncStr [| None]
# - ValStr  [| None]
# - Groups

# Everything in an excel spreadsheet is called an Entity. The root entity is an Excel.
# Inside Excels there are sheets. Inside sheets there are cells. Cells can be Funcs
# or StaticValues or DynamicValues. Inside sheets there can also be plots.

# Each entity can belong to a Group. Groups are sets of entities that share characteristics,
# but in implementation they are treated as traits. An example is a table.

# Functions have codes and function inputs. Function inputs are tree structures that have other entities
# which are (colloquially speaking) the dependencies of the function. Note that Func includes implicit functions
# like 1 + 2 + A1 / (A2 * A3 + A4).

############################################################################ YAML Functionality and Sources

# Yaml Declarations
__SOURCES = "Sources"
__SOURCE = "Source"
__TYPE = "Type"
__PATH = "Path"
__NAME = "Name"         # Optional
__ENTITIES = "Entities" # Optional
# __NAME                # Optional
__LOC = "CellLocation"  # Optional
__CACHE = "CachedVal"   # Optional
__VAL = "FuncVal"       # Optional
__GROUPS = "Groups"     # Optional
__META = "Metadata"     # Optional

# Source Type
__PYTHON_STDOUT_FILE = "PYTHON_STDOUT_FILE"
__EXCEL_FILE = "EXCEL_FILE"

############################################################################ High Level Entities
class Entity():
    # Here are the methods to serialize and derialize from YAML
    # They are re-implemneted in the Excel type
    @staticmethod
    def FromYamlMap(yaml):
        raise NotImplementedError
    @staticmethod
    def ToYamlMap():
        raise NotImplementedError
    
    # Here are the methods to serialize and deserialize from Excel
    # They are re-implemented in the excel type
    @staticmethod
    def ToExcelFile(filepath):
        raise NotImplementedError
    @staticmethod
    def FromExcelFile(filepath):
        raise NotImplementedError
    
    
    # These will be re-implemented by extensions of entity
    # AddToWorkbook is for children of Excel
    def AddToWorkbook(self, workbook):
        raise NotImplementedError
    
    # These two functions are expected to be present in all entities
    def Children(self):
        raise NotImplementedError
    def Type(self):
        raise NotImplementedError
    def Dependable(self):
        return False

class Excel(Entity):
    # TODO FromYamlMap
    # TODO ToYamlMap

    @staticmethod
    def ToExcelFile(filepath):
        wb = openpyxl.Workbook()
        self.AddToWorkbook(wb)
        wb.save(filepath)
    @staticmethod
    def FromExcelFile(filepath):
        # wb = openpyxl.load_all()
        raise NotImplementedError # TODO
        # return Excel() # TODO

    def __init__(self, children):
        self.children = children
    def Children(self):
        return self.children
    def AddToWorkbook(workbook):
        for child in self.children:
            child.AddToWorkbook(workbook)

class Sheet(Entity):
    # TODO
    pass # all the sheets in the excel

# A value should be something that can be read as the input to a function
# (i.e. it is something that is dependable)
class Value(Entity):
    @staticmethod
    def FromStr(string):
        if "(" in string:
            return Func.FromFuncStr(string)
        else:
            return Cell.FromStr(string)
    
    def Dependable(self):
        return True
    pass

class Cell(Value):
    @staticmethod
    def FromStr(string):
        raise NotImplementedError
    @staticmethod
    def ParseCellLoc(locStr):
        # Assume format is col, row like A1 or AB23
        i = 1
        while not locStr[i].isdigit():
            i += 1
        return locStr[:i], int(locStr[i:])
    
    def __init__(self, locStr, deps=None):
        self.col, self.row = Cell.ParseCellLoc(loc)
        self.deps = deps if deps != None else []
        self.id = uuid.uuid4()

# A static value encodes a constant like `hello` or `2`
class StaticValue(Value):
    def __init__(self, val):
        self.val = val

############################################################################ Functions

__ADD = "ADD"
__SUB = "SUB"
__MULT = "MULT"
__DIV = "DIV"
__ADD_DELIM = "+"
__SUB_DELIM = "-"
__MULT_DELIM = "*"
__DIV_DELIM = "/"

__AGG_DELIMS = set([
    __ADD_DELIM,
    __SUB_DELIM,
    __MULT_DELIM,
    __DIV_DELIM,
])
__AGG_DELIM_TO_CODE = {
    __ADD_DELIM: __ADD,
    __SUB_DELIM: __SUB,
    __MULT_DELIM: __MULT,
    __DIV_DELIM: __DIV,
}

# FuncInputs are a tree that encodes the inputs to a function, which may be multiple
# and may have other entities (i.e. Funcs or Cells, etc...).
class FuncInputs():
    @staticmethod
    def FromFuncInputsStr(funcInputsStr):
        inputs = []
        start = 0
        end = 0
        nestLevel = 0
        while end < len(funcInputsStr):
            if funcInputsStr[end] == "(":
                nestLevel += 1
            elif funcInputsStr[end] == ")":
                nestLevel -= 1
            elif funcInputsStr[end] == "," and nestLevel == 0:
                # Remove spaces at the end of the entity pointed to
                realEnd = end - 1
                while funcInputsStr[realEnd] == " ":
                    realEnd -= 1
                # TODO this might not be a function, but instead a cell
                inputs.append(Value.FromStr(funcInputsStr[start : realEnd + 1]))
                # Remove spaces at the start of the next entity
                nextStart = end + 1
                while nextStart < len(funcInputsStr) and funcInputsStr[start] == " ":
                    nextStart += 1
                start = nextStart
                end = nextStart
            elif nestLevel < 0:
                raise Exception("nestLevel is %d".format(nestLevel))
            end += 1
        if nestLevel != 0:
            raise Exception("nestLevel terminated on %d".format(nestLevel))
        return FuncInputs(inputs)
    
    # The list inputs is a list of entities (probably StaticCells, DynamicCells, or Funcs)
    def __init__(self, inputs):
        self.inputs = inputs

    # This should return the entity dependencies (TODO)
    def Deps(self):
        raise NotImplementedError

class Func(Value):
    @staticmethod
    def FromFuncStr(funcStr):
        funcStr = funcStr.strip()
        # i stores the first index after the code/type if this is NOT an implicit/agg function (like +)
        # and the index of the implicit function type (i.e. the agg delimiter) if it's an agg function
        i = 0
        isAgg = False
        isList = False
        while i < len(funcStr) and funcStr[i] != "(":
            if funcStr[i] == ",":
                isList = True
                break
            if funcStr[i] in FuncInputs.__AGG_DELIMS:
                isAgg = True
                break
            i += 1
        if isList:
            raise Exception("List in Func.FromFuncStr: Should be in FuncInputs.FromFuncInputsStr")
        if i == 0:
            # deal with the case (1 + ...)
            isAgg = True
        if not isAgg:
            j = i
            while j < len(funcStr) and funcStr[j] != ")":
                j += 1
            if j != len(funcStr):
                isAgg = True
        # now we recursively parse the inputs of the function, which, if it's an agg
        # are the element before the agg delimiter and after and otherwise is the comma-seperated
        # list between the first and last `(` and `)`
        if isAgg:
            return Func(__AGG_DELIM_TO_CODE[funcStr[i]], Func.FromFuncStr(funcStr[i + 1:]))
        else:
            return Func(funcStr[:i], FuncInputs.FromFuncInputsStr(funcStr[i + 1 : -1]))
    
    def __init__(self, code, funcInputs):
        self.code = code
        self.funcInputs = funcInputs

# A dynamic value encodes an output from a black box
class DynamicValue(Value):
    # TODO
    def Val(self):
        if self.val is None:
            raise NotImplementedError
        else:
            return self.val

class PythonStdout(DynamicValue):
    # TODO
    pass

############################################################################ Parsers

__SRC_REQ = [
    __PATH,
    __TYPE,
]
# Entities have no requirements

__GROUPS = "Groups"    # Optional
__META = "Metadata"    # Optional
__LOC = "CellLocation" # Optional
__CACHE = "CachedVal"  # Optional
__VAL = "FuncVal"      # Optional
__NAME #is optional

# Return a dictionary that returns None for items not in it
class DefaultDict:
    def __init__(self, mp):
        self.mp = mp
    def __getitem__(self, key):
        if item in mp:
            return mp[key]
        else:
            return None
    def __setitem__(self, key, value):
        self.mp[key] = value

# Check that all the required keys are present in the given map and return a DefaultDict of it
# Takes (Name string, Map, Required Keys)
def DefaultRequire(s, mp, rqs):
    for rq in rqs:
        if not rq in mp:
            raise Exception("Malformed Yaml %s: Did not find %s", s, rq)
    return DefaultDict(mp)

def Yaml(filepath):
    with open(filepath) as f:
        return yaml.load_all(f, yaml.loader.SafeLoader)

class Source():
    def __init__(self, path, _type, entities=None, name=None):
        self.id = _id
        self.path = path
        self.type = _type
        self.entities = entities
        self.name = path if name == None else name
    def Type(self):
        self.type

class PythonStdoutFile(Source):
    def __init__(self, path, name=None):
        super.__init__(path, __PYTHON_STDOUT_FILE, entities=None, name=name)
class ExcelFile(Source):
    def __init__(self, path, name=None):
        super.__init(path, __EXCEL_FILE, entities=Excel.FromExcelFile(path), name=name)

# TODO some testing of our object model
if __name__ == "__main__":
    pass