from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE

import pprint as _pp
pp = _pp.PrettyPrinter(indent=2)
def pprint(string):
    pp.pprint(string)

# Yaml Declarations
__SOURCES = "Sources"
__SOURCE = "Source"
__TYPE = "Type"
__PATH = "Path"
__NAME = "Name"         # Optional
__ENTITIES = "Entities" # Optional
_ID = "Id"
# __NAME                # Optional
__LOC = "CellLocation"  # Optional
__CACHE = "CachedVal"   # Optional
__VAL = "FuncVal"       # Optional
__GROUPS = "Groups"     # Optional
__META = "Metadata"     # Optional
__CHILDREN = "Children"

# Source Type
__PYTHON_STDOUT_FILE = "PYTHON_STDOUT_FILE"
__EXCEL_FILE = "EXCEL_FILE"

def deserialize(filename):
    wb = load_workbook(filename=filename)
    yaml = {}
    yaml["sheets"] = []
    for sheet_name in wb.sheetnames:
        sheet = deserialize_sheet(wb[sheet_name])
        sheet["Type"] = "Sheet"
        sheet["Name"] = sheet_name
        yaml["sheets"].append(sheet)
    return yaml

# Note that coords start at 1 instead of 0
# def cell_name2coords(cell_name):
#     pass

# do not support merged cells for now
def deserialize_sheet(sheet):
    yaml = {}
    yaml["cells"] = []
    # Max containing data
    for column in range(sheet.min_column, sheet.max_column + 1):
        for row in range(sheet.min_row, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            if cell.value != None:
                yaml["cells"].append({
                    "Type": "Cell",
                    "Name": cell.coordinate,
                    "FuncValue": cell.value,
                })
    return yaml

if __name__ == "__main__":
    d = deserialize("ex.xlsx")
    print(d)
    # print(dir(d))
    # # print(d.chartsheets)
    # # print(d.get_sheet_names())
    # for sheetName in d.get_sheet_names():
    #     print("Found sheet %s".format(sheetName))
    #     # print()
    #     sheet = d.get_sheet_by_name(sheetName)
    #     print(sheet.calculate_dimension())
    #     cell = sheet.cell(2, 1)
    #     print(cell.internal_value)
    #     print(cell.value)