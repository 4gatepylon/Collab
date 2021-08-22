import re
import yaml
import uuid
import subprocess
from subprocess import Popen, PIPE
from enum import Enum
from glob import glob

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE

# Helpful for debugging
import pprint as _pp
pp = _pp.PrettyPrinter(indent=4)
def pprint(string):
    pp.pprint(string)

# Any process you call in your build cannot take longer than 5 seconds to run
PROC_TIMEOUT = 5

# Yaml Structure is
# A list of sources with a type, path, and entities (if it's an excel file)
__SOURCES = "Sources"
__TYPE = "Type"
__PATH = "Path"
__ENTITIES = "Entities" # Optional

# When you want to have a callname for a file do !callname:<my_call_name>
# Each call name will be able to take in parameters which MUST be cell names or
# COLUMN RANGES. Row ranges are NOT supported. The outputs of functions are placed
# ONE CELL TO THE RIGHT of the text with the callname for visibility. When you have
# multiple ranges, the shortest one is picked, and then iterators go from the beginning of each
# range getting the nth item of each range and passing it in as a parameter to the function.
# BY DEFAULT RANGES ARE CREATED DOWNWARDS.
__CALL_NAME = "CallName"

# Entity Structure
__LOC = "CellLocation"
__VAL = "Value"
__GROUPS = "Groups"

# Deps is a list or dependencies, for a dynamic type it is instead a singleton
# string with the filename of a single source
__DEPS = "Dependencies"

# Source Type
__PYTHON_STDOUT_FILE = "PYTHON_STDOUT_FILE"
__EXCEL_FILE = "EXCEL_FILE"

# Entity Types
__SHEET = "Sheet"
__CELL = "Cell"
__DYNAMIC = "DynamicCell"

# Dynamic deps structure
__STRING = "String"
__CELL_SOURCES = "CellSources"

py_re = re.compile(".*\.py$")
excel_re = re.compile(".*\.xlsx$")

def infer_location(col, row, idx):
   return col + 1, row + idx

call_name_keyword = "callname"
def get_call_name(filepath):
   # print("filepath " + filepath)
   if filepath[-3:] != ".py":
      return None
   with open(filepath) as f:
      l = f.readline().rstrip()
      i = 0
      while i < len(l) and l[i] == "#":
         i += 1
      if i >= len(l):
         return None
      l = l[i+1:].strip()
      if l[0] == "!" or not ":" in l:
         return None
      sp = l.split(":")
      if len(sp) > 2:
         return None
      tk = sp[0].strip().lower()
      c = sp[-1].strip() # call name is case sensitive!
      if tk != call_name_keyword:
         return None
      return c

# Turn Excel into Yaml
def serialize(filenames=None, yaml_path="master.yaml", TESTONLY=False, verbose=False):
   cn2f = {}

   TESTONLY_PREPEND = ""
   if TESTONLY:
      TESTONLY_PREPEND = "TEST_"
   if filenames is None:
      filenames = glob("../*.py") + glob("../*.xlsx")
   serialized = {}
   sources = []
   serialized[__SOURCES] = sources
   # Python files must be done first to get the callnames in there
   for filename in sorted(filenames, key=lambda fn: 0 if ".py" in fn else 1):
      print("Serializing file " + filename)
      matches_py = py_re.match(filename)
      matches_excel = excel_re.match(filename)
      if not matches_py and not matches_excel:
         raise Exception("Cannot parse Non-py or excel file. Found: " + filename)

      entities = []
      source = {
         __TYPE: __EXCEL_FILE if matches_excel else __PYTHON_STDOUT_FILE,
         __PATH: filename,
         __ENTITIES: entities
      }
      if matches_excel:
         wb = load_workbook(filename=filename)
         serialize_entities(entities, wb, cn2f)
      else:
         call_name = get_call_name(filename)
         if not call_name is None:
            if call_name in cn2f:
               raise Exception("Call names must be unique")
            if call_name in FORMULAE:
               raise Exception("Tried to use call name " + call_name + " which is a formulae")
            source[__CALL_NAME] = call_name
            cn2f[call_name] = filename
      sources.append(source)
   try:
      yp, yfp = yaml_path.rsplit("/", 1)
      yp = yp + "/"
   except ValueError:
      yp = ""
      yfp = yaml_path
   with open(yp + TESTONLY_PREPEND + yfp, "w") as yf:
      yaml.dump(serialized, yf)
   if verbose:
      pprint(serialized)

# This will get the arguments of a cell value 
def parse_cn_deps(full_cell_value):
   deps = {__STRING: full_cell_value}
   deps[__CELL_SOURCES] = []
   sp = full_cell_value.split("(")
   if len(sp) > 2:
      raise Exception("Does not support nesting of functions (etc) for call name dependencies")
   if sp[-1][-1] != ")":
      raise Exception("Malformed callname utilization because closing parens is not last char")
   args = sp[-1][: -1].strip()
   if ")" in args:
      raise Exception("Malformed callname utilization because ) somewhere without ( and nesting isn't allowed anyways")
   for agg in ["+", "/", "*", "-"]:
      if agg in args:
         raise Exception("Implicit function (aggregators like +, *, ...) are not allowed in callname arguments")
   if len(args) < 1:
      deps[__CELL_SOURCES].append(None)
      return deps
   # grab the arguments
   tks = list(map(lambda x: x.strip(), args.split(",")))
   # turn them into [(col start, row start), range end (optional)])
   tks = list(map(
      lambda x: list(map(
         lambda y: deserialize_loc(y),
         x.split(":"),
         )),
      tks,
   ))
   # turn the list into one of [[column, row start, number of rows to move forward until first exclusive row], ...]
   tks = list(map(
      lambda tk_ls: simplify_range(tk_ls), 
      tks,
   ))
   if len(tks) < 1:
      raise Exception("WTF")
   # keep adding a tuple of cells to __CELL_SOURCES until we hit the lowest number of rows to move forward
   mdr = tks[0][-1]
   d = 0
   end = False
   while d < mdr:
      this_tup = []
      for c, rs, dr in tks:
         # dr is exclusive
         if d >= dr:
            end = True
            break
         else:
            # example: create [A1, [A1, B3, [A1, B3, C10]
            # then [
            #   [A1, B3, C10],
            #]
            # then [
            # [A1, B3, C10],
            # [A2, B4, C11],
            # [A3, B5, C12],
            # ... 
            #]
            this_tup.append(serialize_loc(c, rs + d))
      if end:
         break
      deps[__CELL_SOURCES].append(this_tup)
      d += 1
   return deps

# Helper for parse_cn_deps
def simplify_range(tk_ls):
   if len(tk_ls) == 1:
      c, r = tk_ls[0]
      return [c, r, 1]
   else:
      if len(tk_ls) != 2:
         raise Exception("Bad format")
      (c, r), (c_end, r_end) = tk_ls
      if c_end != c:
         raise Exception("We only support ranges over columns which means that the data must change row each time, but not column")
      return [c, r, r_end - r + 1]

def is_eq(v):
   return type(v) == str and v[0] == "="

def calc_cell_values(cell_srcs, cell2val):
   vals = []
   for cell_coord in cell_srcs:
      pprint(cell2val)
      v = fetch_or_fail(cell2val, cell_coord)
      if is_eq(v):
         raise Exception("We do not allow reference to non-static cells from python scripts yet")
      vals.append(v)
   return vals

# This is only being called on excep files
def serialize_entities(entities, wb, cn2f):
   for sheet_name in wb.sheetnames:
      # print("Found sheet with name " + sheet_name)
      sheet = wb[sheet_name]
      deps = []
      entities.append({
         __TYPE: __SHEET,
         __VAL: sheet_name,
         __DEPS: deps,
      })
      for column in range(sheet.min_column, sheet.max_column + 1):
         for row in range(sheet.min_row, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            # print(column, row, cell.value is None)
            if not cell.value is None:
               if type(cell.value) == str and "(" in cell.value:
                  cn = cell.value.split("(", 1)[0]
                  # print("hi")
                  if cn in cn2f:
                     # print("byte")
                     print("Calculating dynamic cell " + cell.coordinate + " with val " + str(cell.value))
                     deps.append({
                        __TYPE: __DYNAMIC,
                        __LOC: cell.coordinate,
                        __VAL: cn2f[cn],
                        __DEPS: parse_cn_deps(cell.value),
                     })
                     # Only continue if we find it, else treat it as a regular cell which could be
                     # a function or a string or whatever...
                     continue
               # print("hey", cell.value)
               deps.append({
                  __TYPE: __CELL,
                  __LOC: cell.coordinate,
                  __VAL: cell.value,
               })

# I'm fairly sure this is correct, but I'm too tired to prove it,
# and it's not actually just plain numbering because there is no 0
# i.e. A = 1 not 0, and there does not exist A0, so every time you go up,
# you also shift up...
num_letters = ord("Z") - ord("A") + 1
def letter2number(letter):
   if ord(letter) < ord("A") or ord(letter) > ord("Z"):
      raise Exception("Must be uppercase letter!")
   # A => 1, B => 2, ...
   return ord(letter) - ord("A") + 1
def letters2number(letters):
   # A = 1, B = 2, ...
   # AA = 26 + 1 = 27, BB = 2 * (26) + 2
   acc = 0
   for letter in letters:
      acc *= num_letters
      acc += letter2number(letter)
   return acc
def number2letters(number):
   if number <= 26:
      return chr(ord("A") - 1 + number)
   return number2letters(number // 26) + number2letters(number % 26)
# these two expect/return in the column, row format
def deserialize_loc(loc):
   i = 0
   while i < len(loc) and not loc[i].isdigit():
      i += 1
   return letters2number(loc[:i]), int(loc[i:])
def serialize_loc(col, row):
   return number2letters(col) + str(row)

def fetch_or_fail(maplike, key):
   if not key in maplike:
      raise Exception("Could not find " + key)
   return maplike[key]

def deserialize(yaml_path="master.yaml", TESTONLY=False):
   f2callname = {}

   TESTONLY_PREPEND = ""
   if TESTONLY:
      TESTONLY_PREPEND = "TEST_"
   # Load the YAML that we want to deserialize from
   y = None
   with open(yaml_path) as f:
      y = yaml.load(f, yaml.loader.SafeLoader)
   if y is None:
      raise Exception("Failed to load YAML")
   # Deserialize  from the YAML
   srcs = fetch_or_fail(y, __SOURCES)
   for y_source in srcs:
      fname = fetch_or_fail(y_source, __PATH)
      if excel_re.match(fname):
         wb = Workbook()
         # Quick Hack to remove trash that is automatically generated
         for trash in wb.sheetnames:
            wb.remove(wb[trash])
         # print("sheetnames is " + str(wb.sheetnames))
         ents = fetch_or_fail(y_source, __ENTITIES)
         # Deserialize sheets (only top level entities for now)
         for y_entity in ents:
            t = fetch_or_fail(y_entity, __TYPE)
            if t == __SHEET:
               cell_coord2val = {}
               ws = wb.create_sheet(title=y_entity[__VAL])
               dps = fetch_or_fail(y_entity, __DEPS)
               # Deserialize Cells (only sub top level entities for now)
               # Need to be sorted so that we'll go through all the static cells first
               # NOTE: dependencies from non-python cells on python cells is NOT supported right now
               # (your python cells may depend on numerical constants in excel and you may (visually)
               # depend on python cells, but we do not support more complex relationships yet)
               for y_dep in sorted(dps, key=lambda x: 0 if fetch_or_fail(x, __TYPE) == __CELL else 1):
                  column, row = deserialize_loc(y_dep[__LOC])
                  # Deserialize cells
                  subt = fetch_or_fail(y_dep, __TYPE)
                  if subt == __CELL:
                     d = y_dep[__DEPS] if __DEPS in y_dep else None
                     if not d is None and len(d) > 0:
                        raise Exception("Tracking cross-cell dependencies not yet supported")
                     vv = fetch_or_fail(y_dep, __VAL)
                     _ = ws.cell(column=column, row=row, value=vv)
                     cell_coord2val[serialize_loc(column, row)] = vv
                  ############################################################## Dynamic Deserialization
                  elif subt == __DYNAMIC:
                     # Fetch the file from which to get the output of to input here and the callname
                     file2output = fetch_or_fail(y_dep, __VAL)
                     cn = fetch_or_fail(f2callname, file2output)

                     # Initialize a list of dependencies which is serialize as
                     # {__STRING: callname(arguments), [(cell input tuple), (cell input tuple), ...]}
                     # where each of these tuples can be none if it's inputless
                     deps = y_dep[__DEPS] if __DEPS in y_dep else {__STRING: cn + "()", __CELL_SOURCES : [None]}
                     # rewrite the equation once for visibility
                     _ = ws.cell(column=column, row=row, value=fetch_or_fail(deps, __STRING))

                     # for the entire list of sources create a dependency
                     root_col, root_row = column, row
                     for idx, dep in enumerate(fetch_or_fail(deps, __CELL_SOURCES)):
                        main_params = ""
                        if not dep is None:
                           # main params is a list of raw values and we want to turn it into something that can go
                           # into the inline command that we will run
                           main_params = calc_cell_values(dep, cell_coord2val)
                           # those quotation marks will turn int \"x\" which is what we want to plase INSIDE
                           # the actual main_params value for that string because we want to keep the quotes for the
                           # python file to take em
                           main_params = ",".join(map(lambda x: "\\\"" + x + "\\\"" if type(x) == str else str(x), main_params))
                        
                        # Here is the gimmick we expect
                        folder, pyfile = file2output.rsplit("/", 1)
                        # Create a string for our variable name that they cannot hit
                        nohit = "check_" + str(uuid.uuid4()).replace("-", "_")
                        # Check that it is a valid value to insert into a cell (bool not supported since)
                        # you can create it by passing, say, a positive vs. negative number with an if statement
                        nohit_check = "type(" + nohit + ") == str or type(" + nohit + ") == int or type(" + nohit + ") == float"
                        # Print nohit to standard out so we can catch with subprocess or error out
                        print2stdout = "print(" + nohit + " if " + nohit_check + " else int(None), end=\"\")"
                        cmd = [
                           # Invoke python and run it inline
                           "python3",
                           "-c",
                           # import the file, expecting there to be a main (otherwise it will fail)
                           "from " + pyfile[:-3] + " import main; " + nohit + " = main(" + main_params + "); " + print2stdout,
                        ]
                        # This can raise a TimeoutExpired `https://docs.python.org/3/library/subprocess.html#subprocess.Popen.communicate`
                        outs, errs = Popen(cmd, stdout=PIPE, stderr=PIPE, cwd=folder).communicate(timeout=PROC_TIMEOUT)
                        if errs is None or (len(errs) == 0 and len(outs) > 0):
                           out = None
                           # Quick hack, but not ideal... it may be better to have them specify in the YAML directly
                           try:
                              out = int(outs)
                           except:
                              try:
                                 out = float(outs)
                              except:
                                 out = str(outs)
                           if out is None:
                              raise Exception("Failed to initialize output from stdout")
                           # Store the callname at the infered (downwards) location (yes I know it's bad code)
                           
                           # at a different location store the output
                           column, row = infer_location(root_col, root_row, idx)
                           print("putting " + str(out) + " in " + str(column) + "," + str(row))
                           # inference here is kind of dumb but it's OK
                           c_out = ws.cell(column=column, row=row, value=out)
                           # check the source code at `https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/cell/cell.html#Cell`
                        else:
                           print("WHEN ERR OUTS HAD LEN " + str(len(outs)) + " and were " + str(outs))
                           print("ERR LEN WAS " + str(len(errs)))
                           raise Exception("Failure running subprocess for " + file2output + " with error `" + str(errs) + "`")
                  ############################################################## END (Dynamic Deserialization)
            else:
               raise Exception("Top level entities MUST be sheet, but found type " + t)
         yp = fetch_or_fail(y_source, __PATH)
         try:
            yf, yfl = yp.rsplit("/", 1)
            yf = yf + "/"
         except ValueError:
            yf = ""
            yfl = yp
         wb.save(filename=yf + TESTONLY_PREPEND + yfl)
      # Recall, fname is the name defined above the previous if
      else:
         p = y_source[__PATH] if __PATH in y_source else None
         if py_re.match(fname):
            f2callname[p] = y_source[__CALL_NAME] if __CALL_NAME in y_source else None
         else:
            if p is None:
               p = "Unknown Path"
            raise Exception("Trying to deserialize with unknown file type on path: " + p)


if __name__ == "__main__":
   # These go through master.yaml
   serialize(filenames=None)#=["../fibs.xlsx"] + glob("../*.py"), verbose=True)
   deserialize(TESTONLY=True)
   # pprint(d)
   # print("Serializing into test_output.yaml")
   # serialize_excel_yaml()