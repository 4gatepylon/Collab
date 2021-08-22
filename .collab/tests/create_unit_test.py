import formulas
import openpyxl
import re
import unittest
from collections import namedtuple
from os.path import dirname, abspath
import os

Cell = namedtuple("Cell", "value location")

def create_unittest_from_file(formula, filename):
   matches = re.findall(r'[A-Z]\d',formula)
   wb = openpyxl.load_workbook(filename)
   sheet = wb.active
   values = []
   for match in matches:
      value = sheet[match].value
      values.append(value)

   func = formulas.Parser().ast(formula)[1].compile()

   try:
      result = func(*values)
      return result
   except:
      raise ValueError(f"{formula} is not a valid excel formula")

def create_unit_test_from_formula(formula):
   func = formulas.Parser().ast(formula)[1].compile()
   return func

def retrieve_cell(filename: str, cell: str):
   wb = openpyxl.load_workbook(filename)
   sheet = wb.active
   value = sheet[match].value
   return Cell(value, cell)

def retrieve_col(filename: str, start: str, end: str):
   colname = start[0]
   start_number = int(start[1:])
   end_number = int(end[1:])
   values = []
   wb = openpyxl.load_workbook(filename)
   sheet = wb.active

   for i in range(start_number, end_number+1):
      index = f'{colname}{i}'
      value = sheet[index].value
      values.append(Cell(value, index))

   return values

class TestExcel(unittest.TestCase):
    @classmethod
    def setUp(cls):
      cls.files = {}
      parent_dir = dirname(dirname(dirname(abspath(__file__))))
      for file in os.listdir(f"{parent_dir}/"):
         if file.endswith(".xlsx"):
            cls.files[file] = os.path.join(f"/{parent_dir}", file)
